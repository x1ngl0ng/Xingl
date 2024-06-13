import yaml
import openpyxl
import pymysql

#这个类里的方法，没自测过，可能会报错，主要看下思路
class Login_data:


    Login_data=[
        ["lisi","123456",'个人理财系统']
    ]
    NameErrorData=[
        ["lisi", "12347756", '个人理财系统']
    ]
# excel 以下——————————————————————————————————————————————————————————————————————————————————————————————————————
    def readexcel(self,filename,sheet_name=None):#1个用于读取excel文件的函数，参数代表：要读取的sheet页，默认None
        data = []
        #通过openpyxl模块加载excel，给定路径
        wb = openpyxl.load_workbook(filename=filename)
        #判断参数是不是空值，是的话就第1个
        if sheet_name is None:
            sheet_name = wb.sheetnames[0]  # 默认读取第一个 sheet
        st = wb[sheet_name]#获取到实际的sheet页

        #opebpyxl中的 st.max_row 最大列数，st.max_column最大行数
        for i in range(2, st.max_row + 1):#，左闭右开，从第2行开始遍历到最后1行，
            da = []
            for j in range(1, st.max_column):#从第2列开始遍历到最后1列
                cell_value = st.cell(row=i, column=j).value#获取当前行和列的值
                da.append(cell_value)#把每个单元格中的数据放入da列表
            data.append(da)#再将ta放入data列表
        return data

#以下yaml——————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————
    def readyaml(self,file_path):

        with open(file_path, "r", encoding="utf-8") as file:
            yaml_data = yaml.safe_load(file)

            return yaml_data
#以下db——————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————


    def readmysql(self,host, user, password, sql, params=None, charset = 'utf8mb4', port = '3306'):
        """
        :param host: ……
        :param user: ……
        :param password: ……
        :param sql:
        :param params:sql查询中的参数，给占位符%s实际提供的值
        :param charset:与数据库连接常用的字符集 utf8mb4
        :param port: ……
        :return:
        """

        db_config = {
            'host': host,
            'user': user,
            'password': password,
            # 'database': database,
            'charset': charset,
            'port': port
        }
        connection = None
        result = None

        try:
            connection = pymysql.connect(**db_config)

            with connection.cursor() as cursor:
                cursor.execute(sql, params)

                result = cursor.fetchall()

            connection.commit()

        except Exception as e:
            print(f"Error: {e}")

        finally:
            if connection:
                connection.close()

        return result



if __name__ == '__main__':
    login_data = Login_data()
    login_data.data()
    # data = login_data.loginSuccessDtate()'
    host = 'your_host'
    user = 'your_user'
    password = 'your_password'
    database = 'your_database'
    query = 'SELECT * FROM db.table WHERE column = %s'
    params = ('value',)

    results = readmysql(host, user, password, database, query, params)

    #----------------------------------------------------------------------------------------------
